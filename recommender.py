
import argparse
import re
import numpy as np
import pandas as pd
from rapidfuzz import fuzz, process
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity


RECEPTEN_PAD  = r"recepten_AH.xlsx"
VOORRAAD_PAD  = r"pantry_proxies.xlsx"

GEWICHT_HOEVEELHEID = 0.6
GEWICHT_TFIDF       = 0.4

MIN_ING        = 6      # onder dit aantal ingrediënten gaat de penalty in
PENALTY_KRACHT = 0.4    # maximale penalty (0.4 = max 40% aftrek)

# ── MMR diversiteit ────────────────────────────────────────────────────────────
MMR_LAMBDA = 0.7        # 1.0 = puur score, 0.0 = puur diversiteit

# ── Categorie boost ────────────────────────────────────────────────────────────
MAX_PER_CATEGORIE = 3

# ── Overige instellingen ───────────────────────────────────────────────────────
FUZZY_DREMPEL       = 65
BIJNA_KLAAR_DREMPEL = 0.6

EENHEID_NAAR_BASIS = {
    "kg": 1000, "g": 1, "mg": 0.001, "l": 1000, "ml": 1, "dl": 100,
    "el": 15, "tl": 5, "eetlepel": 15, "theelepel": 5,
    # pantry_proxies gebruikt "stuk" (enkelvoud) ipv "stuks"
    "stuks": 100, "stuk": 100,
    "teen": 5, "takje": 5, "snuf": 2, "snufje": 2,
}

STOP_WOORDEN = {
    "verse", "vers", "gedroogde", "gedroogd", "gerookte", "gerookt", "gesneden",
    "gehakte", "gehakt", "gekookte", "gekookt", "halfvolle", "volle", "magere",
    "milde", "zoete", "zoet", "grote", "groot", "kleine", "klein", "middelgrote",
    "middelgroot", "biologische", "biologisch", "griekse", "ah", "excellent",
    "mix", "voor", "van", "met", "en", "of", "pure", "belegen", "geraspte",
    "lauwwarm", "diepvries", "scharreleieren", "scharrelei",
}

CATEGORIE_KEYWORDS = {
    "ontbijt/brunch": ["wafel", "broodje", "bagel", "ei", "smoothie", "havermout",
                       "toast", "croissant", "brioche", "skyr", "oats", "pancake", "bowl"],
    "soep":           ["soep", "bouillon"],
    "salade":         ["salade", "bowl", "sla"],
    "snack/borrel":   ["snack", "chips", "crisps", "borrel", "amuse", "hoorntje",
                       "dip", "crackers", "chipje", "crostini"],
    "dessert":        ["taart", "cake", "cookie", "muffin", "dessert", "cupcake",
                       "pudding", "crumble", "pavlova", "cheesecake", "tompouce",
                       "rocky road", "melba", "dame blanche"],
    "pasta":          ["pasta", "spaghetti", "penne", "fusilli", "orzo", "lasagne",
                       "tagliatelle", "pappardelle", "parelcouscous"],
    "vlees":          ["kip", "gehakt", "biefstuk", "slavink", "gehaktbal", "burger",
                       "spekjes", "chorizo", "bacon", "lam", "varken"],
    "vis":            ["zalm", "kabeljauw", "forel", "tonijn", "garnalen",
                       "kibbeling", "ansjovis", "makreel"],
    "vegetarisch":    ["tofu", "tempeh", "halloumi", "kikkererwten", "linzen",
                       "falafel", "bloemkool", "aubergine", "courgette"],
    "stamppot":       ["stamppot", "puree", "stoof"],
    "overig":         [],
}


def detecteer_categorie(naam: str) -> str:
    naam_lower = naam.lower()
    for cat, keywords in CATEGORIE_KEYWORDS.items():
        if any(kw in naam_lower for kw in keywords):
            return cat
    return "overig"


# ══════════════════════════════════════════════════════════════════════════════
# Data helpers
# ══════════════════════════════════════════════════════════════════════════════

def normaliseer(tekst: str) -> str:
    tekst  = tekst.lower().strip()
    woorden = [w for w in re.split(r"\s+", tekst)
               if w not in STOP_WOORDEN and len(w) > 2]
    return " ".join(woorden[:3])


def parse_ingredient(tekst: str) -> tuple:
    tekst = tekst.lower().strip()
    m = re.match(r"^([\d.,½¼¾]+)\s*([a-zA-Z]+)?\s+(.+)$", tekst)
    if not m:
        return normaliseer(tekst), 100.0
    hs      = (m.group(1).replace(",", ".").replace("½", "0.5")
               .replace("¼", "0.25").replace("¾", "0.75"))
    eenheid = (m.group(2) or "stuks").lower()
    try:
        h = float(hs)
    except ValueError:
        h = 1.0
    return normaliseer(m.group(3).strip()), h * EENHEID_NAAR_BASIS.get(eenheid, 100)


def bouw_voorraad(voorraad_df: pd.DataFrame, uid: int) -> dict:
    """
    Bouwt een {genormaliseerde_naam: hoeveelheid_in_basisunit} dict
    voor de gegeven gebruiker.

    Aanpassing: kolommen zijn nu lowercase (ingredient / hoeveelheid / eenheid).
    """
    result = {}
    for _, r in voorraad_df[voorraad_df["user_id"] == uid].iterrows():
        naam   = normaliseer(str(r["ingredient"]))        # was: Ingredient
        factor = EENHEID_NAAR_BASIS.get(str(r["eenheid"]).lower(), 100)  # was: Eenheid
        result[naam] = float(r["hoeveelheid"]) * factor   # was: Hoeveelheid
    return result


def fuzzy_match(naam: str, voorraad: dict) -> float:
    if not voorraad or not naam:
        return 0.0
    res = process.extractOne(naam, list(voorraad.keys()), scorer=fuzz.token_set_ratio)
    return voorraad[res[0]] if res and res[1] >= FUZZY_DREMPEL else 0.0


# ══════════════════════════════════════════════════════════════════════════════
# Scoring functies
# ══════════════════════════════════════════════════════════════════════════════

def bereken_hoeveelheid_score(ingredienten: list, voorraad: dict) -> dict:
    if not ingredienten:
        return {"score": 0.0, "gedekt": 0, "gedeeltelijk": 0, "totaal": 0, "details": []}
    dekkingen, details, vol, deel = [], [], 0, 0
    for regel in ingredienten:
        if not regel.strip():
            continue
        naam, benodigd = parse_ingredient(regel)
        beschikbaar    = fuzzy_match(naam, voorraad)
        d = (min(beschikbaar, benodigd) / benodigd
             if benodigd > 0 else (1.0 if beschikbaar > 0 else 0.0))
        dekkingen.append(d)
        details.append({
            "ingredient":  naam,
            "beschikbaar": round(beschikbaar),
            "benodigd":    round(benodigd),
            "dekking":     round(d, 3),
        })
        if d >= 1.0:
            vol  += 1
        elif d > 0:
            deel += 1
    if not dekkingen:
        return {"score": 0.0, "gedekt": 0, "gedeeltelijk": 0, "totaal": 0, "details": []}
    return {
        "score":        round(float(np.mean(dekkingen)), 4),
        "gedekt":       vol,
        "gedeeltelijk": deel,
        "totaal":       len(dekkingen),
        "details":      details,
    }


def lengte_penalty(n_ingredienten: int) -> float:
    """
    Penalty voor recepten met weinig ingrediënten.
    Recept met 1 ingredient → penalty van PENALTY_KRACHT (bijv. 0.4 → score * 0.6)
    Recept met >= MIN_ING   → geen penalty (factor 1.0)
    """
    if n_ingredienten >= MIN_ING:
        return 1.0
    aftrek = PENALTY_KRACHT * (MIN_ING - n_ingredienten) / MIN_ING
    return round(1.0 - aftrek, 4)


def mmr_herschik(
    kandidaten: pd.DataFrame,
    tfidf_vecs,
    recept_idx_map: dict,
    top_n: int,
    lmbda: float = MMR_LAMBDA,
) -> pd.DataFrame:
    """
    Maximal Marginal Relevance herschikking.
    lmbda=1.0 → puur score-gebaseerd
    lmbda=0.0 → puur diversiteit
    """
    geselecteerd = []
    resterende   = kandidaten.copy()

    while len(geselecteerd) < top_n and len(resterende) > 0:
        if len(geselecteerd) == 0:
            beste = resterende.iloc[0]
        else:
            ges_idxs = [recept_idx_map[r["ReceptID"]] for r in geselecteerd]
            ges_vecs = tfidf_vecs[ges_idxs]
            mmr_scores = []
            for _, rij in resterende.iterrows():
                rec_idx = recept_idx_map.get(rij["ReceptID"])
                if rec_idx is None:
                    mmr_scores.append(-999)
                    continue
                rec_vec = tfidf_vecs[rec_idx: rec_idx + 1]
                max_sim = float(cosine_similarity(rec_vec, ges_vecs).max())
                mmr     = lmbda * rij["Score_gecorrigeerd"] - (1 - lmbda) * max_sim
                mmr_scores.append(mmr)
            resterende = resterende.copy()
            resterende["_mmr"] = mmr_scores
            beste = resterende.loc[resterende["_mmr"].idxmax()]

        geselecteerd.append(beste.to_dict())
        resterende = resterende[resterende["ReceptID"] != beste["ReceptID"]]

    result = pd.DataFrame(geselecteerd)
    if "_mmr" in result.columns:
        result = result.drop(columns=["_mmr"])
    return result.reset_index(drop=True)


def categorie_diversiteit(
    df: pd.DataFrame,
    top_n: int,
    max_per_cat: int = MAX_PER_CATEGORIE,
) -> pd.DataFrame:
    """
    Zorgt dat er max MAX_PER_CATEGORIE recepten per categorie in de top-N zitten.
    """
    categorie_tellers = {}
    geselecteerd      = []
    reserve           = []

    for _, rij in df.iterrows():
        cat   = rij.get("Categorie", "overig")
        count = categorie_tellers.get(cat, 0)
        if count < max_per_cat and len(geselecteerd) < top_n:
            geselecteerd.append(rij)
            categorie_tellers[cat] = count + 1
        else:
            reserve.append(rij)

    for rij in reserve:
        if len(geselecteerd) >= top_n:
            break
        geselecteerd.append(rij)

    return pd.DataFrame(geselecteerd).reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# Recommender klasse
# ══════════════════════════════════════════════════════════════════════════════

class WasteLessRecommender:
    def __init__(self, recepten_pad: str, voorraad_pad: str):
        print("Laden data...", end=" ", flush=True)
        self.recepten_df = pd.read_excel(recepten_pad)
        self.voorraad_df = pd.read_excel(voorraad_pad)

        # ── Aanpassing: user_id is integer; gebruikers gesorteerd als int ──────
        self.gebruikers = sorted(self.voorraad_df["user_id"].unique())

        # Voeg categorie en ingrediëntaantal toe
        self.recepten_df["Categorie"] = self.recepten_df["Naam"].apply(detecteer_categorie)
        self.recepten_df["N_ingredienten"] = self.recepten_df["Ingrediënten"].fillna("").apply(
            lambda x: len([r for r in x.splitlines() if r.strip()])
        )

        print(f"✓ ({len(self.recepten_df)} recepten, {len(self.gebruikers)} gebruikers)")
        self._bouw_tfidf()

    def _bouw_tfidf(self):
        print("TF-IDF vectorisatie...", end=" ", flush=True)

        recept_docs = self.recepten_df["Ingrediënten"].fillna("").apply(
            lambda x: " ".join(normaliseer(r) for r in x.splitlines() if r.strip())
        ).tolist()

        # ── Aanpassing: kolom is 'ingredient' (lowercase) ─────────────────────
        # Extra: pantry-categorie als context-woord meegeven aan gebruikersdocument
        # zodat TF-IDF ook op voedselgroep-niveau kan matchen.
        gebruiker_docs = []
        for uid in self.gebruikers:
            subset = self.voorraad_df[self.voorraad_df["user_id"] == uid]
            ing_tokens = " ".join(
                normaliseer(str(r)) for r in subset["ingredient"]
            )
            # Pantry-categorieën als extra tokens (zuivel, vlees_vis, …)
            cat_tokens = " ".join(subset["categorie"].dropna().unique())
            gebruiker_docs.append(f"{ing_tokens} {cat_tokens}")

        self.vectorizer = TfidfVectorizer(ngram_range=(1, 2), min_df=1)
        mat = self.vectorizer.fit_transform(recept_docs + gebruiker_docs)

        self.recept_vecs    = mat[:len(recept_docs)]
        self.gebruiker_vecs = mat[len(recept_docs):]
        self.recept_ids     = self.recepten_df["ID"].tolist()
        self.recept_idx_map = {rid: i for i, rid in enumerate(self.recept_ids)}

        print("✓")

    def aanbeveel(
        self,
        gebruiker_id: int,
        top_n: int = 10,
        alleen_bijna_klaar: bool = False,
        verbose: bool = False,
    ) -> pd.DataFrame:
        # ── Aanpassing: gebruiker_id is int ───────────────────────────────────
        if gebruiker_id not in self.gebruikers:
            raise ValueError(f"Gebruiker '{gebruiker_id}' niet gevonden.")

        uid_idx = list(self.gebruikers).index(gebruiker_id)

        # ── Aanpassing: kolom heet 'profiel' (was: Persona) ───────────────────
        persona  = self.voorraad_df[
            self.voorraad_df["user_id"] == gebruiker_id
        ]["profiel"].iloc[0]

        voorraad = bouw_voorraad(self.voorraad_df, gebruiker_id)

        tfidf_sim = cosine_similarity(
            self.gebruiker_vecs[uid_idx: uid_idx + 1], self.recept_vecs
        )[0]

        kandidaten = []
        for i, (_, recept) in enumerate(self.recepten_df.iterrows()):
            ings  = [r.strip() for r in str(recept.get("Ingrediënten", "")).splitlines() if r.strip()]
            hoev  = bereken_hoeveelheid_score(ings, voorraad)
            n_ing = recept.get("N_ingredienten", len(ings))
            tot, ged = hoev["totaal"], hoev["gedekt"]
            bijna    = bool(tot > 0 and ged / tot >= BIJNA_KLAAR_DREMPEL)

            if alleen_bijna_klaar and not bijna:
                continue

            basis  = round(
                GEWICHT_HOEVEELHEID * hoev["score"] + GEWICHT_TFIDF * float(tfidf_sim[i]),
                4,
            )
            penalty = lengte_penalty(n_ing)
            gecorr  = round(basis * penalty, 4)

            kandidaten.append({
                "ReceptID":             recept.get("ID", ""),
                "Recept":               recept.get("Naam", ""),
                "Categorie":            recept.get("Categorie", "overig"),
                "Score_basis":          basis,
                "Lengte_penalty":       penalty,
                "Score_gecorrigeerd":   gecorr,
                "Hoev_Score":           hoev["score"],
                "TF_IDF_Score":         round(float(tfidf_sim[i]), 4),
                "Gedekt_volledig":      ged,
                "Gedekt_deels":         hoev["gedeeltelijk"],
                "Totaal_ing":           tot,
                "N_ingredienten":       n_ing,
                "Bijna_klaar":          bijna,
                "URL":                  recept.get("URL", ""),
                "_details":             hoev["details"],
            })

        if not kandidaten:
            return pd.DataFrame()

        kand_df = pd.DataFrame(kandidaten).sort_values(
            "Score_gecorrigeerd", ascending=False
        ).reset_index(drop=True)

        pool   = kand_df.head(min(50, len(kand_df)))
        mmr_df = mmr_herschik(
            pool, self.recept_vecs, self.recept_idx_map,
            top_n=min(top_n * 2, len(pool)),
        )
        div_df = categorie_diversiteit(mmr_df, top_n=top_n)
        div_df["Rang"] = div_df.index + 1

        if verbose:
            print(f"\n{'═'*65}")
            print(f"  Top {top_n} voor gebruiker {gebruiker_id} ({persona})")
            print(f"  [penalty min_ing={MIN_ING}, MMR λ={MMR_LAMBDA}, max/cat={MAX_PER_CATEGORIE}]")
            print(f"{'═'*65}")
            for _, r in div_df.iterrows():
                bijna_label = " 🟢 BIJNA KLAAR" if r["Bijna_klaar"] else ""
                pen_label   = f" ⚠ penalty={r['Lengte_penalty']}" if r["Lengte_penalty"] < 1.0 else ""
                print(f"\n  #{r['Rang']:02d}  [{r['Categorie']}] {r['Recept']}{bijna_label}{pen_label}")
                print(f"       Score: {r['Score_gecorrigeerd']:.3f}  "
                      f"(basis: {r['Score_basis']:.3f}, "
                      f"hoeveelheid: {r['Hoev_Score']:.3f}, "
                      f"tfidf: {r['TF_IDF_Score']:.3f})")
                print(f"       {r['Gedekt_volledig']}/{r['Totaal_ing']} volledig, "
                      f"{r['Gedekt_deels']} gedeeltelijk ({r['N_ingredienten']} ing. totaal)")
                for d in r["_details"][:5]:
                    vlag = "✓" if d["dekking"] >= 1 else ("~" if d["dekking"] > 0 else "✗")
                    print(f"         {vlag} {d['ingredient']}: "
                          f"{d['beschikbaar']}/{d['benodigd']} ({d['dekking']:.0%})")
                print(f"       {r['URL']}")
            print()

        return div_df.drop(columns=["_details"])

    def aanbeveel_alle(self, top_n: int = 10) -> pd.DataFrame:
        alle = []
        print(f"\nAanbevelingen voor {len(self.gebruikers)} gebruikers...")
        for i, uid in enumerate(self.gebruikers, 1):
            print(f"  [{i}/{len(self.gebruikers)}] gebruiker {uid}", flush=True)
            df = self.aanbeveel(uid, top_n=top_n)
            if df.empty:
                continue
            # ── Aanpassing: kolom heet 'profiel' ─────────────────────────────
            persona = self.voorraad_df[
                self.voorraad_df["user_id"] == uid
            ]["profiel"].iloc[0]
            df.insert(0, "user_id", uid)
            df.insert(1, "Profiel", persona)    # was: Persona
            alle.append(df)
        return pd.concat(alle, ignore_index=True)

    def exporteer(self, df: pd.DataFrame, pad: str):
        from openpyxl.styles import Alignment, Font, PatternFill

        groen = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        # ── Aanpassing: profielnamen zijn nu lowercase snake_case ─────────────
        profiel_kleuren = {
            "student":      "DDEEFF",
            "alleenstaand": "FFE8CC",
            "gezond_eter":  "DDFFDD",
            "gezin":        "FFEECC",
            "stel":         "FFDDEE",
        }

        with pd.ExcelWriter(pad, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Aanbevelingen")
            ws = writer.sheets["Aanbevelingen"]

            for col, w in zip("ABCDEFGHIJKLMNOP",
                              [12, 14, 5, 14, 45, 12, 12, 12, 12, 14, 16, 14, 12, 12, 12, 55]):
                ws.column_dimensions[col].width = w

            for cel in ws[1]:
                cel.font = Font(bold=True)

            bijna_col = list(df.columns).index("Bijna_klaar") if "Bijna_klaar" in df.columns else -1
            for rij in ws.iter_rows(min_row=2):
                for cel in rij:
                    cel.alignment = Alignment(wrap_text=True, vertical="top")
                if bijna_col >= 0 and rij[bijna_col].value is True:
                    for cel in rij:
                        cel.fill = groen

            if "user_id" in df.columns:
                top1 = df.groupby("user_id").first().reset_index()
                top1.to_excel(writer, index=False, sheet_name="Beste match per gebruiker")
                ws2 = writer.sheets["Beste match per gebruiker"]
                for col, w in zip("ABCDEFGH", [12, 14, 5, 14, 45, 12, 12, 12]):
                    ws2.column_dimensions[col].width = w
                for cel in ws2[1]:
                    cel.font = Font(bold=True)
                # ── Aanpassing: profiel-kolom heet nu 'Profiel' ───────────────
                profiel_col_idx = (list(top1.columns).index("Profiel")
                                   if "Profiel" in top1.columns else 1)
                for rij in ws2.iter_rows(min_row=2):
                    profiel = rij[profiel_col_idx].value if len(rij) > profiel_col_idx else ""
                    kleur   = profiel_kleuren.get(str(profiel).lower(), "FFFFFF")
                    fill    = PatternFill(start_color=kleur, end_color=kleur, fill_type="solid")
                    for cel in rij:
                        cel.fill = fill
                        cel.alignment = Alignment(vertical="top")

            if "Categorie" in df.columns:
                groep_col = "Profiel" if "Profiel" in df.columns else "user_id"
                cat_df = (
                    df.groupby([groep_col, "Categorie"])["Rang"]
                    .count()
                    .reset_index()
                    .rename(columns={"Rang": "Aantal"})
                    .sort_values([groep_col, "Aantal"], ascending=[True, False])
                )
                cat_df.to_excel(writer, index=False, sheet_name="Categorie verdeling")
                ws3 = writer.sheets["Categorie verdeling"]
                for col, w in [("A", 14), ("B", 20), ("C", 10)]:
                    ws3.column_dimensions[col].width = w
                for cel in ws3[1]:
                    cel.font = Font(bold=True)

        print(f"✓ Opgeslagen: {pad}")


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="WasteLess Recommender v2")
    # ── Aanpassing: gebruiker is nu een integer ───────────────────────────────
    parser.add_argument("--gebruiker", type=int,
                        help="Gebruiker ID (integer, bijv. 1 of 42)")
    parser.add_argument("--top",       type=int, default=10)
    parser.add_argument("--alle",      action="store_true")
    parser.add_argument("--bijna",     action="store_true")
    parser.add_argument("--verbose",   action="store_true")
    parser.add_argument("--output",    type=str, default=None)
    args = parser.parse_args()

    rec = WasteLessRecommender(RECEPTEN_PAD, VOORRAAD_PAD)

    if args.alle:
        df  = rec.aanbeveel_alle(top_n=args.top)
        pad = args.output or "aanbevelingen.xlsx"
        rec.exporteer(df, pad)
    elif args.gebruiker is not None:
        df = rec.aanbeveel(
            args.gebruiker,
            top_n=args.top,
            alleen_bijna_klaar=args.bijna,
            verbose=True,
        )
        if args.output:
            rec.exporteer(df, args.output)
    else:
        # Demo: toon top 5 voor drie gebruikers
        for uid in [1, 21, 61]:
            rec.aanbeveel(uid, top_n=5, verbose=True)


if __name__ == "__main__":
    main()