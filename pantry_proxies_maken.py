

import pandas as pd
import random
import re
from collections import defaultdict

random.seed(42)


with open('output.txt', 'r', encoding='utf-8') as f:
    alle_ingredienten = [line.strip() for line in f if line.strip()]


def enkelvoud(s):
    """Eenvoudige Nederlandse enkelvoudsvormen"""
    woorden = s.split()
    if not woorden:
        return s
    last = woorden[-1]
   
    uitzonderingen = {
        'eieren': 'ei', 'uien': 'ui', 'appels': 'appel', 'appelen': 'appel',
        'peren': 'peer', 'citroenen': 'citroen', 'tomaten': 'tomaat',
        'aardappelen': 'aardappel', 'wortelen': 'wortel', 'paprikas': 'paprika',
        'champignons': 'champignon', 'courgettes': 'courgette',
        'kipfilets': 'kipfilet', 'kipdijfilets': 'kipdijfilet',
        'biefstukken': 'biefstuk', 'rookworsten': 'rookworst',
        'scharreleieren': 'scharrelei', 'bosuitjes': 'bosui',
        'krieltjes': 'krieltje', 'blokjes': 'blokje',
    }
    low = last.lower()
    if low in uitzonderingen:
        woorden[-1] = uitzonderingen[low]
    elif low.endswith('jes') and len(low) > 5:
        woorden[-1] = last[:-2]  # tomatenblokjes -> tomatenblokje
    elif low.endswith('tjes') and len(low) > 6:
        woorden[-1] = last[:-3] + 'je'
    return ' '.join(woorden)


seen_norm = {}
for ing in alle_ingredienten:
    norm = enkelvoud(ing).lower().lstrip("'\"\u2018\u2019")
    if norm not in seen_norm:
        seen_norm[norm] = enkelvoud(ing).lstrip("'\"\u2018\u2019")

ingredienten_uniek = list(seen_norm.values())


# (ingredientnaam, eenheid, min_qty, max_qty, kans_aanwezig)
PANTRY = {
    'zuivel': [
        ('ei', 'stuk', 2, 12, 0.90),
        ('scharrelei', 'stuk', 2, 12, 0.60),
        ('melk', 'ml', 250, 1500, 0.85),
        ('halfvolle melk', 'ml', 250, 1000, 0.70),
        ('volle melk', 'ml', 250, 1000, 0.40),
        ('boter', 'g', 50, 250, 0.85),
        ('roomboter', 'g', 50, 250, 0.70),
        ('Griekse yoghurt', 'g', 150, 500, 0.55),
        ('volle yoghurt', 'g', 150, 500, 0.50),
        ('crème fraîche', 'g', 100, 200, 0.45),
        ('slagroom', 'ml', 100, 250, 0.40),
        ('kookroom', 'ml', 100, 200, 0.35),
        ('geraspte kaas', 'g', 50, 200, 0.70),
        ('Parmezaanse kaas', 'g', 30, 150, 0.50),
        ('mozzarella', 'g', 100, 250, 0.45),
        ('feta', 'g', 100, 200, 0.40),
        ('Griekse feta', 'g', 100, 200, 0.35),
        ('ricotta', 'g', 125, 250, 0.30),
        ('cottagecheese', 'g', 150, 300, 0.30),
        ('roomkaas', 'g', 100, 200, 0.40),
        ('mascarpone', 'g', 100, 250, 0.30),
        ('kwark', 'g', 150, 500, 0.45),
    ],
    'groente_vers': [
        ('ui', 'stuk', 1, 5, 0.95),
        ('rode ui', 'stuk', 1, 3, 0.60),
        ('knoflook', 'stuk', 1, 3, 0.90),
        ('tomaat', 'stuk', 1, 6, 0.75),
        ('paprika', 'stuk', 1, 3, 0.65),
        ('rode paprika', 'stuk', 1, 3, 0.55),
        ('wortel', 'stuk', 1, 5, 0.70),
        ('aardappel', 'g', 200, 1000, 0.80),
        ('courgette', 'stuk', 1, 2, 0.50),
        ('champignon', 'g', 100, 400, 0.55),
        ('kastanjechampignon', 'g', 100, 250, 0.40),
        ('spinazie', 'g', 100, 400, 0.50),
        ('sla', 'stuk', 1, 2, 0.45),
        ('rucola', 'g', 50, 150, 0.40),
        ('prei', 'stuk', 1, 2, 0.45),
        ('broccoli', 'stuk', 1, 2, 0.50),
        ('bloemkool', 'stuk', 1, 1, 0.40),
        ('citroen', 'stuk', 1, 4, 0.70),
        ('limoen', 'stuk', 1, 3, 0.45),
        ('verse gember', 'g', 20, 100, 0.50),
        ('cherrytomaat', 'g', 150, 400, 0.50),
        ('avocado', 'stuk', 1, 3, 0.50),
        ('bosui', 'stuk', 2, 6, 0.45),
        ('venkel', 'stuk', 1, 2, 0.30),
        ('zoete aardappel', 'g', 200, 600, 0.40),
        ('pompoen', 'g', 300, 800, 0.30),
        ('asperge', 'g', 200, 500, 0.30),
        ('sperzieboon', 'g', 150, 400, 0.35),
        ('ijsbergsla', 'stuk', 1, 1, 0.40),
        ('komkommer', 'stuk', 1, 2, 0.50),
    ],
    'vlees_vis': [
        ('kipfilet', 'g', 200, 600, 0.70),
        ('kipdijfilet', 'g', 200, 500, 0.55),
        ('kipfiletreepje', 'g', 200, 400, 0.45),
        ('gehakt', 'g', 250, 500, 0.65),
        ('rundergehakt', 'g', 250, 500, 0.55),
        ('half-om-halfgehakt', 'g', 250, 500, 0.50),
        ('zalm', 'g', 150, 400, 0.55),
        ('gerookte zalm', 'g', 100, 200, 0.45),
        ('tonijn in blik', 'g', 100, 200, 0.60),
        ('garnaal', 'g', 150, 300, 0.35),
        ('bacon', 'g', 75, 200, 0.55),
        ('spekreepje', 'g', 75, 150, 0.45),
        ('kipgehakt', 'g', 250, 500, 0.35),
        ('biefstuk', 'stuk', 1, 2, 0.35),
        ('gehaktbal', 'stuk', 4, 12, 0.40),
        ('rookworst', 'stuk', 1, 2, 0.40),
    ],
    'droogwaren': [
        ('pasta', 'g', 250, 500, 0.85),
        ('spaghetti', 'g', 250, 500, 0.75),
        ('penne', 'g', 250, 500, 0.60),
        ('fusilli', 'g', 250, 500, 0.55),
        ('rijst', 'g', 250, 1000, 0.85),
        ('basmatirijst', 'g', 250, 750, 0.60),
        ('bloem', 'g', 100, 1000, 0.75),
        ('zelfrijzend bakmeel', 'g', 100, 500, 0.45),
        ('paneermeel', 'g', 100, 300, 0.40),
        ('havermout', 'g', 100, 500, 0.55),
        ('linze', 'g', 150, 400, 0.45),
        ('kikkererwt', 'g', 150, 400, 0.45),
        ('kidneyboon', 'g', 150, 400, 0.40),
        ('bruine boon', 'g', 150, 400, 0.35),
        ('couscous', 'g', 150, 400, 0.45),
        ('quinoa', 'g', 150, 300, 0.40),
        ('lasagneblad', 'stuk', 6, 12, 0.35),
        ('crackers', 'stuk', 4, 16, 0.50),
        ('broodkruimel', 'g', 50, 200, 0.35),
    ],
    'sauzen_conserven': [
        ('tomatensaus', 'g', 200, 400, 0.70),
        ('passata', 'g', 200, 700, 0.60),
        ('tomatenblokje in blik', 'g', 200, 400, 0.65),
        ('tomatenconcentraat', 'g', 70, 140, 0.60),
        ('kokosmelk', 'ml', 200, 400, 0.55),
        ('bouillonblokje', 'stuk', 2, 6, 0.80),
        ('kippenbouillonblokje', 'stuk', 2, 4, 0.70),
        ('groentebouillonblokje', 'stuk', 2, 4, 0.65),
        ('pesto', 'g', 80, 190, 0.55),
        ('sojasaus', 'ml', 50, 200, 0.65),
        ('ketjap manis', 'ml', 50, 150, 0.50),
        ('Worcestershiresaus', 'ml', 30, 100, 0.40),
        ('mayonaise', 'g', 100, 400, 0.70),
        ('mosterd', 'g', 100, 300, 0.75),
        ('ketchup', 'g', 100, 400, 0.70),
        ('sambal', 'g', 50, 200, 0.55),
        ('sriracha', 'ml', 50, 150, 0.40),
        ('appelciderazijn', 'ml', 50, 250, 0.55),
        ('rode wijnazijn', 'ml', 50, 250, 0.50),
        ('balsamicoazijn', 'ml', 50, 150, 0.55),
        ('honing', 'g', 100, 400, 0.70),
        ('ahornsiroop', 'ml', 50, 250, 0.35),
        ('pindasaus', 'g', 100, 300, 0.45),
        ('tahini', 'g', 100, 300, 0.40),
        ('pindakaas', 'g', 100, 400, 0.65),
    ],
    'olie_vet': [
        ('olijfolie', 'ml', 100, 750, 0.90),
        ('zonnebloemolie', 'ml', 100, 500, 0.70),
        ('olie', 'ml', 100, 500, 0.60),
        ('sesamolie', 'ml', 30, 100, 0.40),
    ],
    'kruiden_specerijen': [
        ('zout', 'g', 50, 500, 0.99),
        ('zwarte peper', 'g', 20, 100, 0.99),
        ('paprikapoeder', 'g', 20, 80, 0.75),
        ('komijn', 'g', 10, 50, 0.60),
        ('koriander', 'g', 10, 50, 0.55),
        ('kurkuma', 'g', 10, 50, 0.55),
        ('kaneel', 'g', 10, 50, 0.65),
        ('nootmuskaat', 'g', 5, 30, 0.60),
        ('oregano', 'g', 5, 30, 0.70),
        ('tijm', 'g', 5, 30, 0.65),
        ('rozemarijn', 'g', 5, 30, 0.55),
        ('basilicum', 'g', 5, 30, 0.65),
        ('knoflookpoeder', 'g', 10, 50, 0.60),
        ('cayennepeper', 'g', 5, 30, 0.55),
        ('chilipoeder', 'g', 5, 30, 0.50),
        ('kerriepoeder', 'g', 10, 50, 0.55),
        ('garam masala', 'g', 10, 40, 0.40),
        ('Italiaanse kruiden', 'g', 10, 40, 0.65),
        ('laurierblad', 'stuk', 2, 8, 0.60),
        ('verse basilicum', 'g', 15, 30, 0.45),
        ('verse peterselie', 'g', 15, 30, 0.50),
        ('verse tijm', 'g', 10, 20, 0.35),
        ('verse koriander', 'g', 15, 30, 0.40),
        ('verse munt', 'g', 10, 20, 0.35),
        ('bieslook', 'g', 10, 20, 0.40),
    ],
    'bakken_zoet': [
        ('suiker', 'g', 100, 1000, 0.80),
        ('basterdsuiker', 'g', 100, 500, 0.55),
        ('poedersuiker', 'g', 50, 300, 0.45),
        ('bruine suiker', 'g', 100, 500, 0.50),
        ('vanillesuiker', 'g', 20, 80, 0.60),
        ('bakpoeder', 'g', 10, 100, 0.65),
        ('baksoda', 'g', 10, 100, 0.45),
        ('cacao', 'g', 50, 200, 0.50),
        ('pure chocolade', 'g', 50, 200, 0.45),
        ('melkchocolade', 'g', 50, 200, 0.40),
        ('rozijn', 'g', 50, 200, 0.45),
        ('amandel', 'g', 50, 200, 0.45),
        ('walnoot', 'g', 50, 200, 0.40),
        ('cashewnoot', 'g', 50, 200, 0.40),
        ('pijnboompit', 'g', 30, 100, 0.40),
        ('gedroogde abrikoos', 'g', 50, 200, 0.30),
        ('kokosrasp', 'g', 50, 150, 0.35),
        ('maizena', 'g', 50, 200, 0.60),
        ('gelatineblad', 'stuk', 2, 8, 0.30),
        ('vanilleextract', 'ml', 10, 50, 0.40),
    ],
    'brood_ontbijt': [
        ('brood', 'stuk', 1, 2, 0.80),
        ('volkoren brood', 'stuk', 1, 1, 0.55),
        ('wrap', 'stuk', 2, 8, 0.55),
        ('tortilla', 'stuk', 2, 8, 0.50),
        ('beschuit', 'stuk', 3, 12, 0.55),
        ('roggebrood', 'stuk', 1, 1, 0.35),
        ('müsli', 'g', 150, 500, 0.45),
        ('granola', 'g', 100, 400, 0.40),
        ('cornflakes', 'g', 100, 400, 0.40),
        ('jam', 'g', 100, 400, 0.65),
        ('appelstroop', 'g', 100, 400, 0.40),
        ('hagelslag', 'g', 100, 400, 0.50),
        ('pindakaas', 'g', 100, 400, 0.65),
    ],
    'divers': [
        ('witte wijn', 'ml', 150, 750, 0.45),
        ('rode wijn', 'ml', 150, 750, 0.45),
        ('bier', 'ml', 330, 660, 0.40),
        ('kippenboullion', 'ml', 200, 1000, 0.40),
        ('groentegroenbouillon', 'ml', 200, 1000, 0.35),
    ]
}


pantry_items = {}
for cat, items in PANTRY.items():
    for naam, eenheid, mn, mx, kans in items:
        pantry_items[naam] = (eenheid, kans, mn, mx, cat)


rows = []
user_profiles = [
    'gezin',        # 25%
    'student',      # 20%
    'stel',         # 25%
    'alleenstaand', # 20%
    'gezond_eter',  # 10%
]
profile_weights = [0.25, 0.20, 0.25, 0.20, 0.10]

for user_id in range(1, 101):
    profiel = random.choices(user_profiles, weights=profile_weights)[0]
    
   
    kans_factor = {
        'gezin': 1.2,
        'student': 0.7,
        'stel': 1.0,
        'alleenstaand': 0.8,
        'gezond_eter': 1.1,
    }[profiel]
    
    for ingredient, (eenheid, kans, mn, mx, categorie) in pantry_items.items():
        aangepaste_kans = min(kans * kans_factor, 0.99)
        if random.random() < aangepaste_kans:
          
            if eenheid in ('g', 'ml'):
                stap = 50 if mx > 200 else 25
                hoeveelheid = round(random.randint(mn, mx) / stap) * stap
                hoeveelheid = max(mn, hoeveelheid)
            else:
                hoeveelheid = random.randint(mn, mx)
            
            rows.append({
                'user_id': user_id,
                'profiel': profiel,
                'categorie': categorie,
                'ingredient': ingredient,
                'hoeveelheid': hoeveelheid,
                'eenheid': eenheid,
            })

df = pd.DataFrame(rows)
print(f"Totaal rijen: {len(df)}")
print(f"Gem. ingrediënten per user: {len(df) / 100:.1f}")
print(df.head(20).to_string())
print("\nPer categorie:")
print(df.groupby('categorie')['ingredient'].count())

df.to_csv('pantry_proxies.csv', index=False)
print("\nOpgeslagen!")
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_csv('pantry_proxies.csv')

wb = Workbook()


ws1 = wb.active
ws1.title = 'Pantry Data'

HEADER_FILL = PatternFill('solid', start_color='2E7D32')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
ROW_ALT_FILL = PatternFill('solid', start_color='F1F8E9')
BORDER = Border(
    bottom=Side(style='thin', color='C8E6C9'),
    right=Side(style='thin', color='C8E6C9'),
)
CAT_COLORS = {
    'zuivel':             'E3F2FD',
    'groente_vers':       'E8F5E9',
    'vlees_vis':          'FCE4EC',
    'droogwaren':         'FFF8E1',
    'sauzen_conserven':   'FFF3E0',
    'olie_vet':           'F3E5F5',
    'kruiden_specerijen': 'E0F7FA',
    'bakken_zoet':        'FBE9E7',
    'brood_ontbijt':      'EFEBE9',
    'divers':             'FAFAFA',
}

headers = ['user_id', 'profiel', 'categorie', 'ingredient', 'hoeveelheid', 'eenheid']
ws1.append(headers)
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')

ws1.row_dimensions[1].height = 22

for row_idx, row in df.iterrows():
    ws1.append(row.tolist())
    excel_row = row_idx + 2
    cat = row['categorie']
    fill_color = CAT_COLORS.get(cat, 'FFFFFF')
    fill = PatternFill('solid', start_color=fill_color)
    for col_idx in range(1, len(headers) + 1):
        cell = ws1.cell(row=excel_row, column=col_idx)
        cell.fill = fill
        cell.font = Font(name='Arial', size=10)
        cell.border = BORDER
        if col_idx in (1, 5):
            cell.alignment = Alignment(horizontal='center')

col_widths = [10, 15, 20, 35, 14, 10]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f"A1:F{len(df)+1}"


ws2 = wb.create_sheet('Samenvatting per User')

summary = df.groupby(['user_id', 'profiel']).agg(
    aantal_ingredienten=('ingredient', 'count'),
    categorieen=('categorie', 'nunique'),
).reset_index()

s_headers = ['user_id', 'profiel', 'aantal_ingredienten', 'categorieen']
ws2.append(s_headers)
for col_idx, h in enumerate(s_headers, 1):
    cell = ws2.cell(row=1, column=col_idx)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 22

PROFIEL_KLEUREN = {
    'gezin':        'BBDEFB',
    'student':      'C8E6C9',
    'stel':         'FFE0B2',
    'alleenstaand': 'F8BBD0',
    'gezond_eter':  'E1BEE7',
}

for row_idx, row in summary.iterrows():
    ws2.append(row.tolist())
    excel_row = row_idx + 2
    kleur = PROFIEL_KLEUREN.get(row['profiel'], 'FFFFFF')
    fill = PatternFill('solid', start_color=kleur)
    for col_idx in range(1, 5):
        cell = ws2.cell(row=excel_row, column=col_idx)
        cell.fill = fill
        cell.font = Font(name='Arial', size=10)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

for i, w in enumerate([10, 15, 22, 15], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'


ws3 = wb.create_sheet('Ingredientfrequentie')

freq = df.groupby(['ingredient', 'categorie'])['user_id'].nunique().reset_index()
freq.columns = ['ingredient', 'categorie', 'aantal_users']
freq = freq.sort_values('aantal_users', ascending=False)

f_headers = ['ingredient', 'categorie', 'aantal_users', '% users']
ws3.append(f_headers)
for col_idx, h in enumerate(f_headers, 1):
    cell = ws3.cell(row=1, column=col_idx)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 22

for row_idx, row in freq.iterrows():
    pct = f"=C{row_idx+2}/100"
    ws3.append([row['ingredient'], row['categorie'], row['aantal_users'], pct])
    excel_row = row_idx + 2
    cat = row['categorie']
    fill = PatternFill('solid', start_color=CAT_COLORS.get(cat, 'FFFFFF'))
    for col_idx in range(1, 5):
        cell = ws3.cell(row=excel_row, column=col_idx)
        cell.fill = fill
        cell.font = Font(name='Arial', size=10)
        cell.border = BORDER
        if col_idx in (3,):
            cell.alignment = Alignment(horizontal='center')
        if col_idx == 4:
            cell.number_format = '0.0%'

for i, w in enumerate([35, 20, 14, 12], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A2'
ws3.auto_filter.ref = f"A1:D{len(freq)+1}"

wb.save('pantry_proxies.xlsx')
print("Opgeslagen!")
